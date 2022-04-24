# -*- coding: utf-8 -*-
# =============================================================================>
# ##############################################################################
# ## 
# ## core.py
# ## 
# ##############################################################################
# =============================================================================>
# imports
import re
import openpyxl
from PIL import Image, ImageDraw, ImageFont
# =============================================================================>
#    
#    Definition
#    
# =============================================================================>
DEFAULT_ROW_HEIGHT = 25.0
DEFAULT_COLUMN_WIDTH = 72.0

MARGIN_BUTTOM = 0.0
MARGIN_TOP = 0.0
MARGIN_LEFT = 0.0
MARGIN_RIGHT = 0.0

MARGIN_BUTTOM_CHAR = 1.0
MARGIN_TOP_CHAR = 1.0
MARGIN_LEFT_CHAR = 1.0
MARGIN_RIGHT_CHAR = 1.0

SCALE_HEIGHT_CHAR = 1.6
SCALE_WIDTH_CHAR = 1.2

# =============================================================================>
#    
#    function
#    
# =============================================================================>
# tools
def w_units_to_pixels(_width):
    _image_font = ImageFont.load_default()
    _tmp = Image.new("RGBA", (1024, 1024), (0, 0, 0, 0))

    _tmp = ImageDraw.Draw(_tmp)
    text_size = _tmp.textsize("a", _image_font)
    text_size = [text_size[0], text_size[1]]

    return text_size[0] * _width

def check_data(checked_data):
    # =========================================================================>
    # 引用  https://www.shibutan-bloomers.com/python_libraly_openpyxl-5/2664/
    # horizontal    : 水平位置の指定(デフォルト:‘general’) オプション定数8種から選択
    # vertical      : 高さ位置の指定(デフォルト:’bottom‘) オプション定数5種から選択
    # text_rotation : 文字の回転角度の指定(デフォルト:0) 0~180まで1度ステップで指定
    # wrap_text     : 折り返し有効(デフォルト:’False’ 無効)
    # shrink_to_fit : 文字の自動サイズ調整(デフォルト:’False’ 無効)
    # indent        : インデントサイズ(デフォルト:0)
    # =========================================================================>
    for _sheet_name in checked_data.keys():
        for i in range(len(checked_data[_sheet_name])):
            if not checked_data[_sheet_name][i][6].shrink_to_fit: # shrink_to_fitがTrueなら計算の必要が無い
                # =============================================================>
                # 文字を取得し，文字切れを判定
                # =============================================================>
                _count_char = 0
                _sum_width = 0
                _sum_height = 0
                
                if not checked_data[_sheet_name][i][7] is None:
                    _char = list(checked_data[_sheet_name][i][7])
                    font = checked_data[_sheet_name][i][5]
                    # =============================================================>
                    while len(_char) > 0:
                        # =========================================================>
                        _cur = get_char_size(_char[0], font)
                        _cur_width = _cur[0]
                        _cur_height = _cur[1]

                        # =========================================================>
                        # 文字にマージンを追加
                        _cur_width, _cur_height = get_char_size_processed(_cur)

                        # =========================================================>
                        # 縦に1つ進む (縦に超えた)で終了
                        if _sum_height + _cur_height < checked_data[_sheet_name][i][4]:
                            # 横方向に進む ((横に超えた and 折り返し設定) or 改行)で折り返し
                            if _sum_width + _cur_width < checked_data[_sheet_name][i][3]:
                                if _char[0] == "\n":
                                    _sum_height += _cur_height
                                    _sum_width = 0
                                    _char.pop(0)
                                else:
                                    _sum_width += _cur_width
                                    _char.pop(0)
                            else:
                                if checked_data[_sheet_name][i][6].wrap_text: # [折り返して全体を表示]設定
                                    # 縦に1つ進む
                                    _sum_height += get_char_size("\n", font)[1]
                                    _sum_width = 0
                                else:
                                    break
                        else:
                            break
                    
                    # =============================================================>
                    # スコアを入力
                    checked_data[_sheet_name][i][8] = len(_char) / len(checked_data[_sheet_name][i][7])

                else:
                    checked_data[_sheet_name][i][8] = 1
    
    return checked_data

def get_char_size(char, font, filename = None):
    """_summary_

    Args:
        char (str): 1 char
        font (object): font object Pillow
        filename (str, optional): if str save char image. Defaults to None.

    Returns:
        _char_size: tuple
    """
    _font_path = get_font_path(font.name)
    _image_font = ImageFont.truetype(_font_path, int(font.size))

    _tmp = Image.new("RGBA", (1, 1), (0, 0, 0, 0))
    _tmp = ImageDraw.Draw(_tmp)

    _char_size = _tmp.textsize(char, _image_font)

    if not filename is None:
        img = Image.new('RGBA', _char_size, (0, 0, 0, 0))
        img_d = ImageDraw.Draw(img)
        img_d.text((0, 0), char, fill = (255,255,255,255), font = _image_font)
        try:
            img.save(filename)
        except Exception as e:
            print(e)

    
    return _char_size

def get_font_path(font_name):
    return "../docs/YuGothR.ttc"

def get_list_index_from_excel(x, y, checked_data_sheet):
    for i in range(len(checked_data_sheet)):
        if x == checked_data_sheet[i][0]:
            if y == checked_data_sheet[i][2]:
                return i
    return None

def get_char_size_margined(char_size):
    char_size = list(char_size)
    char_size[1] += MARGIN_TOP_CHAR 
    char_size[1] += MARGIN_BUTTOM_CHAR
    char_size[0] += MARGIN_LEFT_CHAR 
    char_size[0] += MARGIN_RIGHT_CHAR
    return char_size

def get_char_size_processed(char_size):
    char_size = [char_size[0] * SCALE_WIDTH_CHAR, char_size[1] * SCALE_HEIGHT_CHAR]
    char_size = get_char_size_margined(char_size)
    return char_size

# =============================================================================>
# ways
def load_excel_file(filename):
    """_summary_
    Excelファイルの読み込み

    Args:
        filename (str): ファイル名
    
    Description:
        return dict object {"sheet_name":[[x, x_name, y, width, height, font, alignment, content, indicator], ...], ...}
    
    """
    # =====================================================================>
    _wb = openpyxl.load_workbook(filename)
    _output = {}
    # =====================================================================>

    for i in _wb.sheetnames:
        _sheet_output = []
        _sheet = _wb[i]
        DEFAULT_ROW_HEIGHT = _sheet.sheet_format.defaultRowHeight
        DEFAULT_COLUMN_WIDTH = _sheet.sheet_format.defaultColWidth
        
        for _col in _sheet.columns:
            # =============================================================>
            # get width, height
            _index_col = _col[0].column # y
            _name_col = openpyxl.utils.cell.get_column_letter(_index_col)
            _width = _sheet.column_dimensions[_name_col].width

            if _width is None:
                _width = DEFAULT_COLUMN_WIDTH
            
            for j in _col:
                _index_row = j.row # x
                _height = _sheet.row_dimensions[_index_row].height
                
                if _height is None:
                    _height = DEFAULT_ROW_HEIGHT
                
                # =========================================================>
                # init values
                _content = j.value # inner text
                _font = j.font # font
                _alignment = j.alignment # alignment
                _indicator = 0 # 文字切れ危険度

                _sheet_output.append([
                    _index_col,
                    _name_col,
                    _index_row,
                    w_units_to_pixels(openpyxl.utils.units.points_to_pixels(_width)), openpyxl.utils.units.points_to_pixels(_height), 
                    _font, _alignment, 
                    str(_content), 
                    _indicator
                ])
    
        # =================================================================>
        # delete merged cells
        for j in _sheet.merged_cells.ranges:
            # 結合の左上に全て合計
            _root_idx = None
            _child_idx = None
            for x in range(j.min_col, j.max_col + 1):
                for y in range(j.min_row, j.max_row + 1):
                    if y == j.min_row and x == j.min_col:
                        _root_idx = get_list_index_from_excel(x, y, _sheet_output)
                    else:
                        _child_idx = get_list_index_from_excel(x, y, _sheet_output)
                        if not _root_idx is None:
                            if _sheet_output[_root_idx][0] < _sheet_output[_child_idx][0]:
                                _sheet_output[_root_idx][3] += _sheet_output[_child_idx][3]
                            if _sheet_output[_root_idx][2] < _sheet_output[_child_idx][2]:
                                _sheet_output[_root_idx][4] += _sheet_output[_child_idx][4]
                        
                        _sheet_output.pop(_child_idx)

        _output[i] = _sheet_output

    return _wb, _output

def save_excel_file(raw_workbook, checked_data, filename):
    """_summary_
    Excelファイルの書き込み

    Args:
        raw_workbook (object): raw workbook object
        checked_data (dict): 適用するデータ
        filename (str): ファイル名
    
    Description:
        None
    
    """
    for sheet_name in checked_data.keys():
        for i in checked_data[sheet_name]:
            if i[8] > 0.0:
                bg_color = str.format('FF{:02x}{:02X}{:02X}', 255, 255 - int(255 * i[8]), 255 - int(255 * i[8]))
                my_fill = openpyxl.styles.PatternFill(patternType = 'solid', fgColor = bg_color)
                raw_workbook[sheet_name].cell(row = i[2], column = i[0]).fill = my_fill
    
    raw_workbook.save(filename)

if __name__ == "__main__":
    _wb, _list = load_excel_file("../docs/test1.xlsx")

    _list = check_data(_list)
    
    save_excel_file(_wb, _list, "../docs/output.xlsx")
    
    # =========================================================================>
    # _wb.save("../docs/output.xlsx")
    # =========================================================================>

