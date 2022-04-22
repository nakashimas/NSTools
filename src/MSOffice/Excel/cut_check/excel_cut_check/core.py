# -*- coding: utf-8 -*-
# =============================================================================>
# ##############################################################################
# ## 
# ## core.py
# ## 
# ##############################################################################
# =============================================================================>
# imports
import openpyxl
from PIL import Image, ImageDraw, ImageFont
# =============================================================================>
#    
#    Definition
#    
# =============================================================================>
DEFAULT_ROW_HEIGHT = 15.0
DEFAULT_COLUMN_WIDTH = 10.0
# =============================================================================>
#    
#    function
#    
# =============================================================================>

if __name__ == "__main__":
    # =========================================================================>
    _wb = openpyxl.load_workbook("../docs/test1.xlsx")
    # =========================================================================>
    for i in _wb.sheetnames:
        _sheet = _wb[i]
        DEFAULT_ROW_HEIGHT = _sheet.sheet_format.defaultRowHeight
        DEFAULT_COLUMN_WIDTH = _sheet.sheet_format.defaultColWidth
        
        for _col in _sheet.columns:
            # =================================================================>
            # get width, height
            _index_col = _col[0].column # y
            _name_col = openpyxl.utils.cell.get_column_letter(_index_col)
            _width = _sheet.column_dimensions[_name_col].width

            if _width is None:
                _width = DEFAULT_COLUMN_WIDTH
            
            for j in _col:
                _index_row = j.row # x
                _height = _sheet.row_dimensions[_index_row].height
                
                if _height is None:
                    _height = DEFAULT_ROW_HEIGHT
                
                # =============================================================>
                # 文字を取得し，文字切れを判定
                # =============================================================>
                _content = j.value # inner text
                _font = j.font # font
                _alignment = j.alignment # alignment
                _indicator = 0 # 文字切れ危険度
                # =============================================================>
                # 引用  https://www.shibutan-bloomers.com/python_libraly_openpyxl-5/2664/
                # horizontal    : 水平位置の指定(デフォルト:‘general’) オプション定数8種から選択
                # vertical      : 高さ位置の指定(デフォルト:’bottom‘) オプション定数5種から選択
                # text_rotation : 文字の回転角度の指定(デフォルト:0) 0~180まで1度ステップで指定
                # wrap_text     : 折り返し有効(デフォルト:’False’ 無効)
                # shrink_to_fit : 文字の自動サイズ調整(デフォルト:’False’ 無効)
                # indent        : インデントサイズ(デフォルト:0)
                # =============================================================>

                if not _alignment.shrink_to_fit: # shrink_to_fitがTrueなら計算の必要が無い
                    # =========================================================>
                    # 画像から最大必要サイズを計算
                    pass
                    # =========================================================>
                    # 最大必要サイズとセルのサイズを比較し，文字切れを判定
                
                print(str(_name_col) + str(_index_row), _indicator)
    # =========================================================================>
    _wb.save("../docs/output.xlsx")
    # =========================================================================>

